"""
analyze, output unit info to excel spreadsheets
"""
import typing
import logging
import openpyxl

import sc2_globals
import sc2_data

logger = logging.getLogger(__name__)

tower_column_names = [
    'name',
    'tier',
    'builder',
    'cost',
    'hp',
    'armor-type',
    'damage-type',
    'life-bonus-add',
    'life-bonus-mult',
    'shields',
    'life',
    'life-cost',
    'life-score',
    'life-supply',
    'range',
    'dmg-min',
    'dmg-max',
    'dmg-period',
    'attacks',
    'dps',
    'dps-bonus-add',
    'dps-bonus-mult',
    'wpn-bonus-add',
    'wpn-bonus-mult',
    'wpn-period-mult',
    'energy',
    'dps-cost',
    'dps-supply',
    'dps-score',
    'score',
    'supply',
    'move-speed',
    'abil1',
    'abil2',
    'abil3'
]

send_column_names = [
    'name',
    'score',
    'life-score',
    'dps-score',
    'life',
    'dps',
    'cost',
    'base-cost',
    'income',
    'bounty',
    'life-cost',
    'dps-cost',
    'life-bonus-add',
    'life-bonus-mult',
    'dps-bonus-add',
    'dps-bonus-mult',
    'wpn-bonus-add',
    'wpn-bonus-mult',
    'wpn-period-mult',
    'hp',
    'shields',
    'armor-type',
    'energy',
    'damage-type',
    'range',
    'attacks',
    'dmg-min',
    'dmg-max',
    'dmg-period',
    'move-speed',
    'abil1',
    'abil2',
    'abil3'
]

wave_column_names = [
    'name',
    'wave',
    'count',
    'life',
    'dps',
    'life-total',
    'dps-total',
    'bounty',
    'life-bonus-add',
    'life-bonus-mult',
    'dps-bonus-add',
    'dps-bonus-mult',
    'wpn-bonus-add',
    'wpn-bonus-mult',
    'wpn-period-mult',
    'hp',
    'shields',
    'armor-type',
    'energy',
    'damage-type',
    'range',
    'attacks',
    'dmg-min',
    'dmg-max',
    'dmg-period',
    'move-speed',
    'abil1',
    'abil2',
    'abil3'
]


def index_to_column(index: int) -> str:
    rval = ''
    while True:
        rval = chr(ord('A') + int(index % 26)) + rval
        if index < 26:
            break
        index = index // 26 - 1
    return rval


def get_common_excel_fields(unit: typing.Any, fields: list[str]) -> dict[str, typing.Any]:
    cols = {}
    cols['name'] = unit.name
    cols['life-bonus-add'] = unit.bonuses.life_bonus_add
    cols['life-bonus-mult'] = unit.bonuses.life_bonus_mult
    cols['dps-bonus-add'] = unit.bonuses.dps_bonus_add
    cols['dps-bonus-mult'] = unit.bonuses.dps_bonus_mult
    cols['wpn-bonus-add'] = unit.bonuses.wpn_bonus_add
    cols['wpn-bonus-mult'] = unit.bonuses.wpn_bonus_mult
    cols['wpn-period-mult'] = unit.bonuses.wpn_period_mult
    cols['hp'] = unit.hp
    cols['shields'] = unit.shields
    cols['armor-type'] = unit.armor
    cols['energy'] = unit.energy
    cols['damage-type'] = unit.wpn.damage.dmg_type
    cols['range'] = unit.wpn.range
    cols['attacks'] = unit.wpn.targets
    cols['dmg-min'] = unit.wpn.damage.min
    cols['dmg-max'] = unit.wpn.damage.max
    cols['dmg-period'] = unit.wpn.period
    cols['move-speed'] = unit.move_speed
    cols['dps'] = (
        '={dps-bonus-mult}{row}*({dps-bonus-add}{row}+{attacks}{row}*{wpn-bonus-mult}{row}*'
        '(({dmg-min}{row}+{dmg-max}{row})/2+{wpn-bonus-add}{row})/'
        '({dmg-period}{row}/{wpn-period-mult}{row}))'
    )
    cols['life'] = '={life-bonus-mult}{row}*({hp}{row}+{shields}{row}+{life-bonus-add}{row})'
    if 'cost' in fields:
        cols['dps-cost'] = '={dps}{row}/{cost}{row}'
        cols['dps-score'] = '={dps-cost}{row}/AVERAGE({dps-cost}2:{dps-cost}65535)'
        cols['life-cost'] = '={life}{row}/{cost}{row}'
        cols['life-score'] = '={life-cost}{row}/AVERAGE({life-cost}2:{life-cost}65535)'
        cols['score'] = '={dps-score}{row}*{life-score}{row}'
    if 'supply' in fields:
        cols['dps-supply'] = '={dps}{row}/{supply}{row}'
        cols['life-supply'] = '={life}{row}/{supply}{row}'
    return cols


def get_unit_abils(tower: typing.Any, abils_data: dict[str, sc2_data.Ability], max_abilities: int = 3) -> list[str]:
    abil_ids = tower.abilities[:max_abilities]
    abils = [abils_data[a].desc for a in abil_ids]
    abils.extend([''] * max(0, max_abilities - len(abils)))
    return abils


def output_tower_statistics(towers: dict[str, sc2_data.Tower], abils: dict[str, sc2_data.Ability]) -> None:
    # update headers
    # turns dict into str with
    units_fields = []
    for tower in towers.values():
        fields: dict[str, typing.Any] = {}
        fields['cost'] = tower.cost.minerals
        fields['supply'] = tower.cost.supply * (-1)
        fields['tier'] = tower.tier
        fields['builder'] = tower.builder
        fields['abil1'], fields['abil2'], fields['abil3'] = get_unit_abils(tower, abils)
        fields.update(get_common_excel_fields(tower, tower_column_names))
        units_fields.append(fields)
    output_units_excel('towers.xlsx', tower_column_names, units_fields)


def output_send_statistics(sends: dict[str, sc2_data.Send], abils: dict[str, sc2_data.Ability]) -> None:
    sends_fields = []
    for send in sends.values():
        fields: dict[str, typing.Any] = {}
        fields['base-cost'] = send.cost.vespene
        fields['income'] = send.income
        fields['cost'] = fields['base-cost'] + (fields['base-cost'] - 20 * fields['income'])
        fields['bounty'] = send.bounty
        fields['abil1'], fields['abil2'], fields['abil3'] = get_unit_abils(send, abils)
        fields.update(get_common_excel_fields(send, send_column_names))
        sends_fields.append(fields)
    output_units_excel('sends.xlsx', send_column_names, sends_fields)


def output_wave_statistics(waves: dict[str, sc2_data.Wave], abils: dict[str, sc2_data.Ability]) -> None:
    waves_fields = []
    for wave in waves.values():
        fields: dict[str, typing.Any] = {}
        fields['wave'] = wave.index
        fields['count'] = wave.count
        fields['bonus'] = wave.get_wave_bonus()
        fields['abil1'], fields['abil2'], fields['abil3'] = get_unit_abils(wave, abils)
        fields['bounty'] = wave.bounty
        fields.update(get_common_excel_fields(wave, wave_column_names))
        waves_fields.append(fields)
        fields['dps-total'] = '={dps}{row}*{count}{row}'
        fields['life-total'] = '={life}{row}*{count}{row}'
    output_units_excel('waves.xlsx', wave_column_names, waves_fields)


def worksheets_equal(new: openpyxl.Workbook, old: openpyxl.Workbook) -> bool:
    new_sheet = new.worksheets[0]
    old_sheet = old.worksheets[0]
    if new_sheet.dimensions != old_sheet.dimensions:
        return False
    for row in new_sheet.iter_rows():
        for cell in row:
            cell_new = cell.value
            cell_loc = cell.coordinate
            cell_old = old_sheet[cell_loc].value
            if cell_new != cell_old:
                if cell_new in (None, '') and cell_old in (None, ''):
                    continue
                return False
    return True


def output_units_excel(filename: str, fields: list[str], units: list[dict[str, typing.Any]]) -> None:
    """
    args:
    - filename: name of the file (not path, we are keeping all stats in same dir)
    - fields: header basically. List of column names (sorted)
    - cols: dict of unit ids with values of all columns in row as dict(key: col name)
    """
    wb = openpyxl.Workbook()
    sheet = wb.worksheets[0]
    sheet.title = 'Units'

    filepath = sc2_globals.PROJECT_DIR / f'output/spreadsheets/{sc2_data.get_version()}/{filename}'
    sheet.append(fields)
    field_ids = {field: index_to_column(cnum) for cnum, field in enumerate(fields)}

    for row, unit in enumerate(units, start=2):
        unit['dps'] = unit['dps'].format(row=row, **field_ids)
        unit['life'] = unit['life'].format(row=row, **field_ids)
        if 'supply' in fields:
            unit['dps-supply'] = unit['dps-supply'].format(row=row, **field_ids)
            unit['life-supply'] = unit['life-supply'].format(row=row, **field_ids)
        if 'cost' in fields:
            unit['dps-cost'] = unit['dps-cost'].format(row=row, **field_ids)
            unit['life-cost'] = unit['life-cost'].format(row=row, **field_ids)
            unit['dps-score'] = unit['dps-score'].format(row=row, **field_ids)
            unit['life-score'] = unit['life-score'].format(row=row, **field_ids)
            unit['score'] = unit['score'].format(row=row, **field_ids)
        if 'count' in fields:
            unit['dps-total'] = unit['dps-total'].format(row=row, **field_ids)
            unit['life-total'] = unit['life-total'].format(row=row, **field_ids)
        sheet.append({field_ids[field]: unit[field] for field in fields})

    header_style = openpyxl.styles.NamedStyle(name='header')
    header_style.font = openpyxl.styles.Font(bold=True)
    header_style.alignment = openpyxl.styles.Alignment(vertical='center')
    header_style.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'))
    sheet.freeze_panes = 'B2'
    for cell in sheet[1]:
        cell.style = header_style
    try:
        try:
            old = openpyxl.load_workbook(filepath)
        except FileNotFoundError:
            logger.info('Spreadsheet doesn\'t exist, creating %s', filepath)
            wb.save(filepath)
            return
        if not worksheets_equal(wb, old):
            logger.info('Creating %s', filepath)
            wb.save(filepath)
    except PermissionError as e:
        logger.error('Failed to output %s, permission denied %s', filename, str(e))
    except OSError as e:
        logger.error('Failed to output %s, unexpected error occurred %s', filename, str(e))


def output_unit_statistics(towers: dict[str, sc2_data.Tower], sends: dict[str, sc2_data.Send], waves: dict[str, sc2_data.Wave], abils: dict[str, sc2_data.Ability]) -> None:
    dir_path = sc2_globals.PROJECT_DIR / f'output/spreadsheets/{sc2_data.get_version()}'
    if not dir_path.exists():
        dir_path.mkdir(parents=True)
        logger.info('Creating %s dir..', dir_path)
    output_tower_statistics(towers, abils)
    output_send_statistics(sends, abils)
    output_wave_statistics(waves, abils)
